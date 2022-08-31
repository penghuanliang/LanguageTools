import openpyxl
import os
from pathlib import Path
from lxml import etree

from xml.dom import minidom


class XLSXToXML:
    list_name = []

    list_all = []

    def dict_to_xml(self, dir_path, data):
        # dir_path = "C:/Users/Administrator/Desktop/res/"
        file_path = ""
        root = etree.Element("resources")
        for key, value in data.items():
            if value is None:
                continue

            if key == "filename":
                file_path = f"{dir_path}/{value}/strings.xml"
                continue
            else:
                if key == 'string-array':
                    error = etree.SubElement(root, 'string-array', attrib={'name': 'error_codes'})
                    for index in range(0, len(value)):
                        e = etree.SubElement(error, 'item')
                        e.text = value.get(index)
                else:
                    tag = etree.SubElement(root, "string", attrib={'name': key})
                    tag.text = value

        # 判断文件夹是否存在，不存在则
        file = Path(dir_path)
        if file.exists() is False:
            os.mkdir(dir_path)

        file_path_dir = Path(file_path).parent
        if file_path_dir.exists() is False:
            os.mkdir(file_path_dir)

        # soup = BeautifulSoup(etree.tostring(root), 'xml')

        xml_str = minidom.parseString(etree.tostring(root)).toprettyxml(indent='    ')
        xml_str = xml_str.replace('<?xml version="1.0" ?>', '<?xml version="1.0" encoding="UTF-8"?>')
        # print(xml_str)
        # print(soup.prettify())
        with open(file_path, 'wb') as file:
            file.write(xml_str.encode('utf-8'))
            file.close()

    def parse_string(self, worksheet):
        # 1. Get name by first column.
        for column in worksheet.iter_cols(max_col=1, min_row=2):
            for cell in column:
                self.list_name.append(cell.value)

        # 2.Use the
        for column in worksheet.iter_cols(min_col=2):
            dict_wrap = {

            }
            for index in range(0, len(column)):
                cell = column[index]
                if 0 == index:
                    dict_wrap['filename'] = cell.value
                else:
                    dict_wrap[self.list_name[index - 1]] = cell.value
            self.list_all.append(dict_wrap)

    def query_dict_wrap(self, file_name):
        for item_value in self.list_all:
            if item_value['filename'] == file_name:
                return item_value

    def parse_string_array(self, worksheet):
        for column in worksheet.iter_cols(min_col=3):
            item_dict = {}
            string_array = {}
            for index in range(0, len(column)):
                cell = column[index]
                if 0 == index:
                    c = self.query_dict_wrap(cell.value)
                    if c is None:
                        continue
                    item_dict = c
                else:
                    cell_name = worksheet[f'B{index + 1}'].value
                    string_array[cell_name] = cell.value
            item_dict['string-array'] = string_array

    def run(self, _in, _out):
        wb = openpyxl.open(_in)
        self.parse_string(wb.worksheets[0])
        self.parse_string_array(wb.worksheets[1])

        for item in self.list_all:
            self.dict_to_xml(_out, item)
