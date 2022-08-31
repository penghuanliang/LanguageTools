import time, openpyxl, os
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.styles import Font


class NoTranslate:
    list_string = []
    list_string_array = []
    base_name = []

    # Set string-array resource.
    font = Font(
        name='Arial',
        size=10
    )

    def load_file(self, file_dir):
        string_files = []

        file_dir_path = Path(file_dir)
        if file_dir_path.exists() is False:
            return
        files = os.listdir(file_dir)
        for f in files:
            f_strings = Path(f'{file_dir}/{f}/strings.xml')
            if f_strings.exists() is True:
                string_files.append(f_strings)

        for f in string_files:
            string_container = {}
            string_array_container = {}
            name = f.parent.name

            string_container['filename'] = name
            string_array_container['filename'] = name
            soup = BeautifulSoup(f.read_text('utf-8'), 'xml')
            tags = soup.findAll('string')
            for tag in tags:
                attr_name = tag['name']
                if name == 'values':
                    if 0 == len(self.base_name):
                        self.base_name.append('stringName')
                    self.base_name.append(attr_name)
                string_container[attr_name] = tag.string
            self.list_string.append(string_container)

            tags = soup.findAll('item')
            for index in range(0, len(tags)):
                string_array_container[index] = tags[index].string
            self.list_string_array.append(string_array_container)

        # print(list_string)
        # print(list_string_array)
        # print(base_name)

    def generate_xsl(self, in_file, filename):
        self.load_file(in_file)

        for column_index in range(0, len(self.list_string) - 1):
            # 每次遍历创建一个excel表
            wb = openpyxl.Workbook()
            # Delete default worksheet as the workbook was created with at least one worksheets.
            wb.remove(wb['Sheet'])
            sheet_first = wb.create_sheet('string')

            base = self.list_string[0]
            cont = self.list_string[column_index + 1]

            xsls_name = cont['filename']

            write_index = 1

            for row_index in range(0, len(self.base_name)):
                name = self.base_name[row_index]
                if 0 == row_index:
                    sheet_first.cell(write_index, 1, name)
                    sheet_first.cell(row_index + 1, 2, base['filename'])
                    sheet_first.cell(row_index + 1, 3, cont['filename'])
                else:
                    value = cont.get(name)
                    if value is None:
                        write_index += 1
                        sheet_first.cell(write_index, 1, name)
                        sheet_first.cell(write_index, 2, base[name])
                    else:
                        continue

            for col_cell in sheet_first.iter_cols():
                for cell in col_cell:
                    cell.font = self.font

            file = Path(filename, xsls_name + '.xlsx')
            if file.exists() is False:
                open(file, mode='w')
            wb.save(file)

            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())), '>>>>', xsls_name, '\t', '未翻译文本导出完毕')

        print('Completed！！！')
