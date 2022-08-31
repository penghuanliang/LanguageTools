import openpyxl
import os
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill


class GenerateXLSX:
    list_string = []
    list_string_array = []
    base_name = []

    def load_file(self, file_dir):
        string_files = []

        file_dir_path = Path(file_dir)
        if file_dir_path.exists() is False:
            return
        files = os.listdir(file_dir)
        for f in files:
            f_strings = Path(f'{file_dir}/{f}/strings.xml')
            if f_strings.exists() is True:
                string_files.append(f_strings)

        for f in string_files:
            string_container = {}
            string_array_container = {}
            name = f.parent.name

            string_container['filename'] = name
            string_array_container['filename'] = name
            soup = BeautifulSoup(f.read_text('utf-8'), 'xml')
            tags = soup.findAll('string')
            for tag in tags:
                attr_name = tag['name']
                if name == 'values':
                    if 0 == len(self.base_name):
                        self.base_name.append('stringName')
                    self.base_name.append(attr_name)
                string_container[attr_name] = tag.string
            self.list_string.append(string_container)

            tags = soup.findAll('item')
            for index in range(0, len(tags)):
                string_array_container[index] = tags[index].string
            self.list_string_array.append(string_array_container)

        # print(list_string)
        # print(list_string_array)
        # print(base_name)

    def generate_xsl(self, in_file, filename):
        self.load_file(in_file)

        wb = openpyxl.Workbook()
        # Delete default worksheet as the workbook was created with at least one worksheets.
        wb.remove(wb['Sheet'])
        sheet_first = wb.create_sheet('string')
        sheet_second = wb.create_sheet('string-array')

        for row_index in range(0, len(self.base_name)):

            name = self.base_name[row_index]
            sheet_first.cell(row_index + 1, 1, name)

            for column_index in range(0, len(self.list_string)):
                d = self.list_string[column_index]
                if 0 == row_index:
                    sheet_first.cell(row_index + 1, column_index + 2, d['filename'])
                else:
                    cell_value = d.get(name)
                    if cell_value is None:
                        red_style = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        sheet_first.cell(row_index + 1, column_index + 2).fill = red_style
                        continue
                    sheet_first.cell(row_index + 1, column_index + 2, cell_value)

        # Set string-array resource.
        length = len(self.list_string_array[0])
        for row_index in range(0, length):
            if 0 == row_index:
                sheet_second.cell(row_index + 1, 1, 'arrayName')
                sheet_second.cell(row_index + 1, 2, 'itemIndex')
            else:
                sheet_second.cell(row_index + 1, 1, 'error_codes')
                sheet_second.cell(row_index + 1, 2, row_index - 1)

            for column_index in range(0, len(self.list_string_array)):
                d = self.list_string_array[column_index]
                if 0 == row_index:
                    sheet_second.cell(row_index + 1, column_index + 3, d['filename'])
                else:
                    cell_value = d.get(row_index - 1)
                    if cell_value is None:
                        sheet_first.cell(row_index + 1, column_index + 2).fill = PatternFill(start_color='FFFF0000',
                                                                                             end_color='FFFF0000',
                                                                                             fill_type='solid')
                        continue
                    sheet_second.cell(row_index + 1, column_index + 3, cell_value)

        font = Font(
            name='Arial',
            size=10
        )

        for col_cell in sheet_first.iter_cols():
            for cell in col_cell:
                cell.font = font

        for col_cell in sheet_second.iter_cols():
            for cell in col_cell:
                cell.font = font

        wb.save(filename)
        print('翻译文本导出完毕')
